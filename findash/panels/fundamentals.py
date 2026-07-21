"""Financials panel — Bloomberg FA style. Income / Balance / Cash Flow
statements, annual or quarterly, in a flipped table (line items as rows,
periods as columns)."""

from __future__ import annotations

import webbrowser
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from PySide6.QtCore import Qt
from PySide6.QtGui import QAction, QColor
from PySide6.QtWidgets import (
    QApplication,
    QButtonGroup,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QMenu,
    QPushButton,
    QTableWidgetItem,
)

from ..components import MarketTable, make_filter_edit
from ..panel import Panel, register_panel
from ..theme import DOWN

STATEMENTS = [("income", "Income"), ("balance", "Balance"), ("cashflow", "Cash Flow")]
PERIODS = [("annual", "Annual"), ("quarterly", "Quarterly")]

# Yahoo Finance's statement-specific URL slug, keyed by our internal ids.
_YF_SLUG = {"income": "financials", "balance": "balance-sheet", "cashflow": "cash-flow"}


def _fmt_compact(value: Any) -> str:
    """Human-format a financial-statement value: T/B/M suffixes, plain for
    small magnitudes, negatives keep their sign."""
    if value is None:
        return "-"
    try:
        v = float(value)
    except (TypeError, ValueError):
        return "-"
    if v != v:  # NaN (pandas turns None into NaN)
        return "-"
    sign = "-" if v < 0 else ""
    av = abs(v)
    for suffix, div in (("T", 1e12), ("B", 1e9), ("M", 1e6)):
        if av >= div:
            return f"{sign}{av / div:.1f}{suffix}"
    if av >= 1e3:
        return f"{sign}{av:,.0f}"
    return f"{sign}{av:,.2f}"


@register_panel(id="fundamentals", title="Financials", category="Research")
class FundamentalsPanel(Panel):
    def build(self) -> None:
        self._statement = "income"
        self._period = "annual"
        self._data: dict = {}

        # -- statement picker + period toggle -------------------------------
        picker_row = QHBoxLayout()
        self._statement_buttons: dict[str, QPushButton] = {}
        stmt_group = QButtonGroup(self)
        stmt_group.setExclusive(True)
        for key, label in STATEMENTS:
            btn = QPushButton(label, self)
            btn.setCheckable(True)
            btn.clicked.connect(lambda _=False, k=key: self._set_statement(k))
            stmt_group.addButton(btn)
            picker_row.addWidget(btn)
            self._statement_buttons[key] = btn
        picker_row.addSpacing(16)

        self._period_buttons: dict[str, QPushButton] = {}
        period_group = QButtonGroup(self)
        period_group.setExclusive(True)
        for key, label in PERIODS:
            btn = QPushButton(label, self)
            btn.setCheckable(True)
            btn.clicked.connect(lambda _=False, k=key: self._set_period(k))
            period_group.addButton(btn)
            picker_row.addWidget(btn)
            self._period_buttons[key] = btn
        picker_row.addStretch(1)

        self._browser_btn = QPushButton("Browser", self)
        self._browser_btn.setToolTip(
            "Open the company's SEC filings (EDGAR) or this statement on"
            " Yahoo Finance"
        )
        browser_menu = QMenu(self._browser_btn)
        edgar_act = QAction("Company filings (SEC EDGAR)", browser_menu)
        edgar_act.triggered.connect(self._open_filings)
        browser_menu.addAction(edgar_act)
        yahoo_act = QAction("Statement on Yahoo Finance", browser_menu)
        yahoo_act.triggered.connect(self._open_in_browser)
        browser_menu.addAction(yahoo_act)
        self._browser_btn.setMenu(browser_menu)
        picker_row.addWidget(self._browser_btn)

        self._download_btn = QPushButton("Download", self)
        self._download_btn.setToolTip("Export the displayed table (CSV / Excel)")
        self._download_btn.clicked.connect(self._download)
        picker_row.addWidget(self._download_btn)

        self._copy_btn = QPushButton("Copy", self)
        self._copy_btn.setToolTip("Copy the displayed table to the clipboard")
        self._copy_btn.clicked.connect(self._copy_to_clipboard)
        picker_row.addWidget(self._copy_btn)

        self.content_layout.addLayout(picker_row)
        self._update_buttons()
        self._update_actions()

        # -- table -----------------------------------------------------------
        self.table = MarketTable(0, 1, self)
        self.table.setHorizontalHeaderLabels(["Line Item"])
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch
        )

        self._filter = make_filter_edit(self.table, "Find line item…")
        self.content_layout.addWidget(self._filter)
        self.content_layout.addWidget(self.table, 1)

    # -- statement / period toggles -------------------------------------------

    def _update_buttons(self) -> None:
        for key, btn in self._statement_buttons.items():
            btn.setChecked(key == self._statement)
        for key, btn in self._period_buttons.items():
            btn.setChecked(key == self._period)

    def _set_statement(self, key: str) -> None:
        if key == self._statement:
            return
        self._statement = key
        self._update_buttons()
        self._render()

    def _set_period(self, key: str) -> None:
        if key == self._period:
            return
        self._period = key
        self._update_buttons()
        self._render()

    # -- linked-symbol lifecycle ------------------------------------------------

    def on_symbol(self, symbol: str) -> None:
        self.set_status(f"{symbol} loading…")
        self._data = {}
        self.unsubscribe_all()
        self.subscribe(f"financials:{symbol}", self._on_financials)
        self._update_actions()

    def _on_financials(self, data: Any) -> None:
        self._data = data if isinstance(data, dict) else {}
        self._render()

    # -- rendering -------------------------------------------------------------

    def _current_block(self) -> Optional[dict]:
        stmt = self._data.get(self._statement)
        if not isinstance(stmt, dict):
            return None
        block = stmt.get(self._period)
        return block if isinstance(block, dict) else None

    def _render(self) -> None:
        block = self._current_block()
        columns = block.get("columns") if block else None
        rows = block.get("rows") if block else None
        columns = columns if isinstance(columns, list) else []
        rows = rows if isinstance(rows, list) else []

        headers = ["Line Item"] + [str(c) for c in columns]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(0)

        for row_data in rows:
            if not isinstance(row_data, (list, tuple)) or not row_data:
                continue
            label = row_data[0]
            values = list(row_data[1:])
            r = self.table.rowCount()
            self.table.insertRow(r)
            label_item = QTableWidgetItem(str(label) if label is not None else "-")
            label_item.setFlags(label_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.table.setItem(r, 0, label_item)
            for col in range(len(columns)):
                value = values[col] if col < len(values) else None
                item = QTableWidgetItem(_fmt_compact(value))
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
                )
                try:
                    if value is not None and float(value) < 0:
                        item.setForeground(QColor(DOWN))
                except (TypeError, ValueError):
                    pass
                self.table.setItem(r, col + 1, item)

        self.table.apply_filter(self._filter.text())

        sym = self.current_symbol or "—"
        stmt_label = dict(STATEMENTS).get(self._statement, self._statement)
        period_label = dict(PERIODS).get(self._period, self._period)
        self.set_status(f"{sym} · {stmt_label} · {period_label} · {len(rows)} lines")
        self._update_actions()

    # -- statement actions: browser / download / copy -------------------------

    def _update_actions(self) -> None:
        """Enable/disable the browser/download/copy actions to match whether
        a symbol is linked and whether the displayed statement has rows."""
        has_symbol = bool(self.current_symbol)
        block = self._current_block()
        has_rows = bool(block and block.get("rows"))
        self._browser_btn.setEnabled(has_symbol)
        self._download_btn.setEnabled(has_rows)
        self._copy_btn.setEnabled(has_rows)

    def _open_in_browser(self) -> None:
        sym = self.current_symbol
        if not sym:
            return
        slug = _YF_SLUG.get(self._statement, "financials")
        webbrowser.open(f"https://finance.yahoo.com/quote/{sym}/{slug}")

    def _open_filings(self) -> None:
        """The company's actual filings on SEC EDGAR, pre-filtered to 10-K
        (annual) or 10-Q (quarterly). EDGAR resolves plain US tickers with
        dots swapped for dashes (BRK.B -> BRK-B); exchange-suffixed foreign
        listings won't match — the Yahoo action covers those."""
        sym = self.current_symbol
        if not sym:
            return
        cik = sym.replace(".", "-")
        form = "10-K" if self._period == "annual" else "10-Q"
        webbrowser.open(
            "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany"
            f"&CIK={cik}&type={form}&dateb=&owner=include&count=40"
        )

    def _current_dataframe(self) -> Optional[pd.DataFrame]:
        """The currently displayed statement/period as a DataFrame, columns
        as periods, plus a leading "Line Item" column — same shape as the
        on-screen table."""
        block = self._current_block()
        if not block:
            return None
        columns = block.get("columns")
        rows = block.get("rows")
        columns = columns if isinstance(columns, list) else []
        rows = rows if isinstance(rows, list) else []
        if not rows:
            return None

        labels = []
        values_by_col: list[list] = [[] for _ in columns]
        for row_data in rows:
            if not isinstance(row_data, (list, tuple)) or not row_data:
                continue
            labels.append(row_data[0])
            values = list(row_data[1:])
            for col in range(len(columns)):
                values_by_col[col].append(values[col] if col < len(values) else None)

        df = pd.DataFrame({str(c): v for c, v in zip(columns, values_by_col)})
        df.insert(0, "Line Item", labels)
        return df

    def _default_filename(self, ext: str) -> str:
        sym = self.current_symbol or "symbol"
        return f"{sym}_{self._statement}_{self._period}{ext}"

    def export_to_path(self, path: str) -> bool:
        """Write the currently displayed table to ``path`` (.csv or .xlsx).
        Returns True on success; reports write errors via ``set_status``
        instead of raising. Factored out of ``_download`` so it can be
        exercised directly, without a file dialog."""
        df = self._current_dataframe()
        if df is None:
            self.set_status("⚠ nothing to export")
            return False
        try:
            if path.lower().endswith(".xlsx"):
                self._write_xlsx(path, df)
            else:
                # CSV carries no styling, so raw floats read terribly — write
                # the values exactly as the panel displays them (391.0B, 6.11).
                # Full-precision numbers live in the .xlsx export.
                self._display_frame(df).to_csv(path, index=False)
        except Exception as exc:
            self.set_status(f"⚠ {exc}")
            return False
        return True

    def _display_frame(self, df: pd.DataFrame) -> pd.DataFrame:
        """Copy of the export frame with value columns rendered the way the
        on-screen table shows them (compact T/B/M figures)."""
        out = df.copy()
        for col in out.columns[1:]:
            out[col] = [_fmt_compact(v) for v in out[col]]
        return out

    def _write_xlsx(self, path: str, df: pd.DataFrame) -> None:
        """Styled Excel export: title bar, dark header band, thousands-
        separated figures with red negatives, frozen panes, sized columns."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        stmt_label = dict(STATEMENTS).get(self._statement, self._statement)
        period_label = dict(PERIODS).get(self._period, self._period)
        sym = self.current_symbol or "—"

        wb = Workbook()
        ws = wb.active
        ws.title = f"{sym} {stmt_label}"[:31]
        n_cols = len(df.columns)

        # row 1: title band across the full table
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title = ws.cell(row=1, column=1, value=f"{sym} — {stmt_label} ({period_label})")
        title.font = Font(bold=True, size=13, color="FFFFFF")
        title.fill = PatternFill("solid", fgColor="1B2530")
        title.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 24

        # row 2: column headers (Line Item + one column per period)
        header_fill = PatternFill("solid", fgColor="1A2129")
        header_border = Border(bottom=Side(style="medium", color="FFAB2E"))
        for col, name in enumerate(df.columns, start=1):
            cell = ws.cell(row=2, column=col, value=str(name))
            cell.font = Font(bold=True, size=10, color="D7DDE3")
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = Alignment(
                horizontal="left" if col == 1 else "right", vertical="center"
            )
        ws.row_dimensions[2].height = 18

        # data: label column left, figures right with thousands separators;
        # small magnitudes (per-share items) keep decimals, negatives go red
        zebra = PatternFill("solid", fgColor="F3F5F7")
        for r, row in enumerate(df.itertuples(index=False), start=3):
            for c, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                if r % 2 == 1:
                    cell.fill = zebra
                if c == 1:
                    cell.font = Font(size=10)
                    continue
                cell.alignment = Alignment(horizontal="right")
                cell.font = Font(size=10)
                if isinstance(value, (int, float)):
                    fmt = "#,##0" if abs(value) >= 1000 else "#,##0.00"
                    cell.number_format = f"{fmt};[Red]({fmt})"

        label_width = max(
            [len(str(v)) for v in df.iloc[:, 0]] + [12]
        )
        ws.column_dimensions["A"].width = min(label_width + 2, 48)
        for col in range(2, n_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16
        ws.freeze_panes = "B3"
        wb.save(path)

    def _download(self) -> None:
        if self._current_dataframe() is None:
            return
        fn, _ = QFileDialog.getSaveFileName(
            self,
            "Export financials",
            self._default_filename(".csv"),
            "CSV (*.csv);;Excel (*.xlsx)",
        )
        if not fn:
            return
        if self.export_to_path(fn):
            self.set_status(f"exported {Path(fn).name}")

    def clipboard_text(self) -> str:
        """Tab-separated version of the currently displayed table."""
        df = self._current_dataframe()
        if df is None:
            return ""
        return df.to_csv(sep="\t", index=False)

    def _copy_to_clipboard(self) -> None:
        text = self.clipboard_text()
        if not text:
            self.set_status("⚠ nothing to copy")
            return
        QApplication.clipboard().setText(text)
        self.set_status("copied")

    # -- persistence -------------------------------------------------------------

    def settings(self) -> dict:
        return {"statement": self._statement, "period": self._period}

    def restore(self, settings: dict) -> None:
        if not isinstance(settings, dict):
            return
        stmt = settings.get("statement")
        if stmt in self._statement_buttons:
            self._statement = stmt
        period = settings.get("period")
        if period in self._period_buttons:
            self._period = period
        self._update_buttons()
        self._render()
